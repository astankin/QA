import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, col_num).value


def writeData(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, col_num).value = data
    workbook.save(file)


def fillGreenColor(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(row_num, col_num).fill = green_fill
    workbook.save(file)


def fillRedColor(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(row_num, col_num).fill = red_fill
    workbook.save(file)
